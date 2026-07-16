"""
Step 4: Converts the updated database back into Excel/CSV lot files, preserving the original layout.
Using the original file as a base, it overwrites the FL/LL cells with the updated values from the database.
Only rewrites the FL/LL cells that have non-blank values in the database. All other cells are preserved.

How to use:
    python 04_export_lot.py lot1.xlsx                       (export a single lot file)
    python 04_export_lot.py lot1.xlsx lot2.xlsx lot5.csv    (export multiple lot files)
    python 04_export_lot.py --all                           (export all lot files in the database)
"""


import csv
import os
import sqlite3
import sys
from openpyxl import load_workbook

LOTS_FOLDER = "data/lots"
DB_PATH = "database/lot_data.db"
OUTPUT_FOLDER = "output/integrated"

# Determines if a value is considered blank (empty, None, or "nan").
def is_blank(val):
    return val is None or str(val).strip() == "" or str(val).strip().lower() == "nan"

# Retrieves the updated FL/LL values for a given lot file from the database, along with their column positions.
def get_lot_updates(conn, lot_file):
    cur = conn.cursor()
    cur.execute(
        "SELECT fl_col, ll_col FROM lot_metadata WHERE lot_file = ?",
        (lot_file,),
    )
    row = cur.fetchone()
    if row is None:
        return None, None, None
    fl_col, ll_col = row

    cur.execute(
        "SELECT row_index, fl, ll FROM testing_details WHERE lot_file = ?",
        (lot_file,),
    )
    updates = {r[0]: (r[1], r[2]) for r in cur.fetchall()}
    return updates, fl_col, ll_col

# Updates the FL/LL values in an Excel lot file based on the updates, preserving the original layout.
def export_xlsx(lot_file, updates, fl_col, ll_col):
    src_path = os.path.join(LOTS_FOLDER, lot_file)
    wb = load_workbook(src_path)
    ws = wb.active

    updated_cells = 0
    for row_index, (fl, ll) in updates.items():
        excel_row = row_index + 1  # note: openpyxl is 1-indexed, our row_index is 0-indexed
        if not is_blank(fl):
            ws.cell(row=excel_row, column=fl_col + 1).value = fl
            updated_cells += 1
        if not is_blank(ll):
            ws.cell(row=excel_row, column=ll_col + 1).value = ll
            updated_cells += 1

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    out_path = os.path.join(OUTPUT_FOLDER, lot_file)
    wb.save(out_path)
    return out_path, updated_cells

# Updates the FL/LL values in a CSV lot file based on the updates, preserving the original layout.
def export_csv(lot_file, updates, fl_col, ll_col):
    src_path = os.path.join(LOTS_FOLDER, lot_file)
    with open(src_path, newline="", encoding="utf-8-sig", errors="replace") as f:
        rows = [row for row in csv.reader(f)]

    updated_cells = 0
    for row_index, (fl, ll) in updates.items():
        if row_index >= len(rows):
            continue
        row = rows[row_index]
        max_needed = max(fl_col, ll_col)
        while len(row) <= max_needed:
            row.append("")
        if not is_blank(fl):
            row[fl_col] = fl
            updated_cells += 1
        if not is_blank(ll):
            row[ll_col] = ll
            updated_cells += 1

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    out_path = os.path.join(OUTPUT_FOLDER, lot_file)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    return out_path, updated_cells


def export_one(conn, lot_file):
    updates, fl_col, ll_col = get_lot_updates(conn, lot_file)
    if updates is None:
        print(f"  SKIPPED: {lot_file} -- not found in database")
        return

    ext = os.path.splitext(lot_file)[1].lower()
    if ext in {".xlsx", ".xls"}:
        out_path, updated_cells = export_xlsx(lot_file, updates, fl_col, ll_col)
    elif ext == ".csv":
        out_path, updated_cells = export_csv(lot_file, updates, fl_col, ll_col)
    else:
        print(f"  SKIPPED: {lot_file} -- unsupported extension")
        return

    print(f"  OK: {lot_file} -- {updated_cells} cells written -> {out_path}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python 04_export_lot.py <lot_file.xlsx> [more files...]")
        print("       python 04_export_lot.py --all")
        return

    conn = sqlite3.connect(DB_PATH)

    if sys.argv[1] == "--all":
        cur = conn.cursor()
        cur.execute("SELECT lot_file FROM lot_metadata ORDER BY lot_file")
        targets = [row[0] for row in cur.fetchall()]
    else:
        targets = sys.argv[1:]

    print(f"Exporting {len(targets)} file(s)...\n")
    for lot_file in targets:
        export_one(conn, lot_file)

    conn.close()


if __name__ == "__main__":
    main()