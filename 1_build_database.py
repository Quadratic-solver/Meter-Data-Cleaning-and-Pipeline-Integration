"""
Step 1: Build the SQLite database and migrate all 68 lot files into it.

Run this ONCE (or whenever the raw lot files change). After this,
all matching/updating logic can happen against the .db file instead
of re-opening 68 Excel/CSV files every time.

Usage (from VS Code terminal, inside your project folder):
    python 01_build_database.py
"""

import csv
import glob
import os
import re
import sqlite3
from openpyxl import load_workbook

# Configuration
LOTS_FOLDER = "data/lots"
DB_PATH = "database/lot_data.db"
SUPPORTED_PATTERNS = ["*.csv", "*.xlsx", "*.xls"]


# Functions for reading and normalizing lot files and extracting metadata

# convert a header cell to lowercase text form
def normalize_header(cell):
    if cell is None:
        return ""
    text = str(cell).strip().lower()
    return re.sub(r"[\s ]+", " ", text)

# nomralizing serial values by removing spaces and making them uppercase
def normalize_serial(val):
    if val is None:
        return ""
    text = str(val).strip()
    if text.lower() == "nan":
        return ""
    return re.sub(r"\s+", "", text).upper()

# Scans the raw grid to find serial, FL, and LL columns.
def find_header_positions(rows):
    for r_idx, row in enumerate(rows):
        for c_idx, cell in enumerate(row):
            if "serial" in normalize_header(cell):
                serial_col_idx = c_idx
                fl_col_idx = ll_col_idx = None

                for cc_idx, cc in enumerate(row):
                    token = normalize_header(cc)
                    if token == "fl":
                        fl_col_idx = cc_idx
                    elif token == "ll":
                        ll_col_idx = cc_idx

                data_start_idx = r_idx + 1
                if (fl_col_idx is None or ll_col_idx is None) and r_idx + 1 < len(rows):
                    sub_row = rows[r_idx + 1]
                    for cc_idx, cc in enumerate(sub_row):
                        token = normalize_header(cc)
                        if token == "fl":
                            fl_col_idx = cc_idx
                        elif token == "ll":
                            ll_col_idx = cc_idx
                    if fl_col_idx is not None and ll_col_idx is not None:
                        data_start_idx = r_idx + 2

                if fl_col_idx is not None and ll_col_idx is not None:
                    return r_idx, serial_col_idx, fl_col_idx, ll_col_idx, data_start_idx
    return None

# Finda the metadata row (BRAND/TYPE/WIRE/...) and the values row right below it.
def find_metadata_row(rows):
    for r_idx, row in enumerate(rows):
        tokens = [normalize_header(c) for c in row]
        if "brand" in tokens and "type" in tokens:
            values_row = rows[r_idx + 1] if r_idx + 1 < len(rows) else []
            return {normalize_header(c): (values_row[i] if i < len(values_row) else None)
                    for i, c in enumerate(row) if c is not None}
    return {}


# ---- Schema ----

SCHEMA = """
CREATE TABLE IF NOT EXISTS lot_metadata (
    lot_file TEXT PRIMARY KEY,
    brand TEXT,
    type TEXT,
    wire TEXT,
    phase TEXT,
    amperes TEXT,
    voltage TEXT,
    kh TEXT,
    dial_constant TEXT,
    header_row_idx INTEGER,
    serial_col INTEGER,
    fl_col INTEGER,
    ll_col INTEGER,
    data_start_idx INTEGER,
    file_type TEXT
);

CREATE TABLE IF NOT EXISTS testing_details (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    lot_file TEXT NOT NULL,
    row_index INTEGER NOT NULL,
    serial TEXT,
    serial_normalized TEXT,
    fl TEXT,
    ll TEXT,
    raw_row TEXT
);

CREATE INDEX IF NOT EXISTS idx_serial_norm ON testing_details(serial_normalized);
CREATE INDEX IF NOT EXISTS idx_lot_file ON testing_details(lot_file);
"""


def get_lot_files():
    files = []
    for pattern in SUPPORTED_PATTERNS:
        files.extend(glob.glob(os.path.join(LOTS_FOLDER, pattern)))
    return sorted(files)


def read_csv_rows(file_path):
    with open(file_path, newline="", encoding="utf-8-sig", errors="replace") as f:
        return [row for row in csv.reader(f)]


def read_xlsx_rows(file_path):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows

# Reads a lot file and inserts its rows into the SQLite database.
def migrate_file(conn, file_path):
    file_name = os.path.basename(file_path)
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        rows = read_csv_rows(file_path)
        file_type = "csv"
    elif ext in {".xlsx", ".xls"}:
        rows = read_xlsx_rows(file_path)
        file_type = "xlsx"
    else:
        print(f"  SKIPPED: {file_name} -- unsupported extension {ext}")
        return False

    positions = find_header_positions(rows)
    if positions is None:
        print(f"  SKIPPED: {file_name} -- could not locate serial/FL/LL columns")
        return False

    header_row_idx, serial_col, fl_col, ll_col, data_start = positions
    meta = find_metadata_row(rows)

    cur = conn.cursor()
    cur.execute(
        """INSERT OR REPLACE INTO lot_metadata
           (lot_file, brand, type, wire, phase, amperes, voltage, kh, dial_constant,
            header_row_idx, serial_col, fl_col, ll_col, data_start_idx, file_type)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            file_name,
            meta.get("brand"), meta.get("type"), meta.get("wire"), meta.get("phase"),
            meta.get("amperes"), meta.get("voltage"), meta.get("kh"), meta.get("dial constant"),
            header_row_idx, serial_col, fl_col, ll_col, data_start, file_type,
        ),
    )

    # clear any old rows for this file before re-inserting (safe re-runs)
    cur.execute("DELETE FROM testing_details WHERE lot_file = ?", (file_name,))

    rows_to_insert = []
    for r_idx in range(data_start, len(rows)):
        row = rows[r_idx]
        max_needed_col = max(serial_col, fl_col, ll_col)
        if len(row) <= max_needed_col:
            continue

        serial_raw = row[serial_col]
        serial_norm = normalize_serial(serial_raw)
        if not serial_norm:
            continue

        rows_to_insert.append((
            file_name,
            r_idx,
            serial_raw,
            serial_norm,
            row[fl_col],
            row[ll_col],
            repr(row),  # keep full original row as a safety net / for debugging
        ))

    cur.executemany(
        """INSERT INTO testing_details
           (lot_file, row_index, serial, serial_normalized, fl, ll, raw_row)
           VALUES (?,?,?,?,?,?,?)""",
        rows_to_insert,
    )
    conn.commit()

    print(f"  OK: {file_name} -- {len(rows_to_insert)} rows migrated "
          f"(header_row={header_row_idx}, serial_col={serial_col}, fl_col={fl_col}, ll_col={ll_col})")
    return True


def main():
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(SCHEMA)

    lot_files = get_lot_files()
    print(f"Found {len(lot_files)} files in {LOTS_FOLDER}\n")

    migrated = 0
    skipped = 0
    for file_path in lot_files:
        try:
            ok = migrate_file(conn, file_path)
            migrated += 1 if ok else 0
            skipped += 0 if ok else 1
        except Exception as e:
            print(f"  FAILED: {os.path.basename(file_path)} -- {e}")
            skipped += 1

    conn.close()

    print("\n=== SUMMARY ===")
    print(f"Files migrated: {migrated}")
    print(f"Files skipped: {skipped}")
    print(f"Database saved to: {DB_PATH}")


if __name__ == "__main__":
    main()