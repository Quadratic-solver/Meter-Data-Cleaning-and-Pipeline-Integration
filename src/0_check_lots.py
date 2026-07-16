"""
Step 0: This will check all lot files from the data gathered before transferring to the database

This script will check if the .xlsx/.xls/.csv files are fit for the mirgation
it will catch defoormed and broken files and will list them in a report. So 
you will know which files are healthy before running the migration.
It will not touch the database or write anything, it's a dry-run check.

WARNING: This pipeline processes internal meter-testing data. Output files
may contain sensitive serial numbers and test values.

Bash:
    python 0_check_lots.py
"""

import csv
import glob
import os
import re
from openpyxl import load_workbook

LOTS_FOLDER = "data/lots"
REPORT_PATH = "database/lot_validation_report.csv"
SUPPORTED_PATTERNS = ["*.csv", "*.xlsx", "*.xls"]

# converts header cells to a simple lowercase text form
def normalize_header(cell):
    if cell is None:
        return ""
    text = str(cell).strip().lower()
    return re.sub(r"[\s ]+", " ", text)

# find the columns for serial, FL, and LL headers and their respective rows
def find_header_positions(rows):
    for r_idx, row in enumerate(rows):
        for c_idx, cell in enumerate(row):
            if "serial" in normalize_header(cell):
                serial_col_idx = c_idx
                fl_col_idx = ll_col_idx = None

                for cc_idx, cc in enumerate(row):
                    token = normalize_header(cc)
                    if token == "fl":
                        fl_col_idx = cc_idx
                    elif token == "ll":
                        ll_col_idx = cc_idx

                data_start_idx = r_idx + 1
                if (fl_col_idx is None or ll_col_idx is None) and r_idx + 1 < len(rows):
                    sub_row = rows[r_idx + 1]
                    for cc_idx, cc in enumerate(sub_row):
                        token = normalize_header(cc)
                        if token == "fl":
                            fl_col_idx = cc_idx
                        elif token == "ll":
                            ll_col_idx = cc_idx
                    if fl_col_idx is not None and ll_col_idx is not None:
                        data_start_idx = r_idx + 2

                if fl_col_idx is not None and ll_col_idx is not None:
                    return r_idx, serial_col_idx, fl_col_idx, ll_col_idx, data_start_idx
    return None

# Collect the lot files from the data/lots folder.
def get_lot_files():
    files = []
    for pattern in SUPPORTED_PATTERNS:
        files.extend(glob.glob(os.path.join(LOTS_FOLDER, pattern)))
    return sorted(files)


def read_csv_rows(file_path):
    with open(file_path, newline="", encoding="utf-8-sig", errors="replace") as f:
        return [row for row in csv.reader(f)]


def read_xlsx_rows(file_path):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows

# Check if the lot file is readable and has a usable header/data structure.
def validate_file(file_path):
    file_name = os.path.basename(file_path)
    ext = os.path.splitext(file_path)[1].lower()

    result = {
        "file": file_name,
        "status": None,
        "reason": "",
        "total_rows": 0,
        "data_rows_found": 0,
        "empty_serial_rows": 0,
        "header_row": None,
        "serial_col": None,
        "fl_col": None,
        "ll_col": None,
    }

    # 1. Checks if the file can be opened and read into rows
    try:
        if ext == ".csv":
            rows = read_csv_rows(file_path)
        elif ext in {".xlsx", ".xls"}:
            rows = read_xlsx_rows(file_path)
        else:
            result["status"] = "FAIL"
            result["reason"] = f"unsupported extension {ext}"
            return result
    except Exception as e:
        result["status"] = "FAIL"
        result["reason"] = f"could not open file: {e}"
        return result

    result["total_rows"] = len(rows)

    if len(rows) == 0:
        result["status"] = "FAIL"
        result["reason"] = "file is empty"
        return result

    # 2. Check if headers can be found for serial, FL, and LL columns
    positions = find_header_positions(rows)
    if positions is None:
        result["status"] = "FAIL"
        result["reason"] = "could not locate serial/FL/LL header columns"
        return result

    header_row_idx, serial_col, fl_col, ll_col, data_start = positions
    result["header_row"] = header_row_idx
    result["serial_col"] = serial_col
    result["fl_col"] = fl_col
    result["ll_col"] = ll_col

    # 3. Checks if there are any usable data rows
    data_rows_found = 0
    empty_serial_rows = 0
    max_needed_col = max(serial_col, fl_col, ll_col)

    for r_idx in range(data_start, len(rows)):
        row = rows[r_idx]
        if len(row) <= max_needed_col:
            continue
        serial_val = row[serial_col]
        if serial_val is None or str(serial_val).strip() == "":
            empty_serial_rows += 1
            continue
        data_rows_found += 1

    result["data_rows_found"] = data_rows_found
    result["empty_serial_rows"] = empty_serial_rows

    if data_rows_found == 0:
        result["status"] = "WARN"
        result["reason"] = "headers found but zero usable data rows"
    else:
        result["status"] = "PASS"
        result["reason"] = ""

    return result


def main():
    os.makedirs("database", exist_ok=True)
    lot_files = get_lot_files()
    print(f"Found {len(lot_files)} files in {LOTS_FOLDER}\n")

    results = []
    for file_path in lot_files:
        r = validate_file(file_path)
        results.append(r)

        if r["status"] == "PASS":
            print(f"  PASS  {r['file']:<20} -- {r['data_rows_found']} data rows "
                  f"(header_row={r['header_row']}, serial_col={r['serial_col']}, "
                  f"fl_col={r['fl_col']}, ll_col={r['ll_col']})")
        elif r["status"] == "WARN":
            print(f"  WARN  {r['file']:<20} -- {r['reason']}")
        else:
            print(f"  FAIL  {r['file']:<20} -- {r['reason']}")

    # write a CSV report
    with open(REPORT_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(results[0].keys()))
        writer.writeheader()
        writer.writerows(results)

    passed = sum(1 for r in results if r["status"] == "PASS")
    warned = sum(1 for r in results if r["status"] == "WARN")
    failed = sum(1 for r in results if r["status"] == "FAIL")

    print("\n=== SUMMARY ===")
    print(f"PASS: {passed}")
    print(f"WARN (headers ok, no usable rows): {warned}")
    print(f"FAIL (could not process): {failed}")
    print(f"\nFull report saved to: {REPORT_PATH}")

    if failed > 0:
        print("\nFiles that FAILED need manual inspection before migration:")
        for r in results:
            if r["status"] == "FAIL":
                print(f"  - {r['file']}: {r['reason']}")


if __name__ == "__main__":
    main()
